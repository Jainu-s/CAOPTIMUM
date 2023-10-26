
import os
import re
from openpyxl import load_workbook
from rapidfuzz import fuzz
from CAOPTIMUM.Config.config import extract_text_from_image
from nltk.corpus import stopwords


'''This code performs test case validation by comparing expected results in Excel files with
 actual results extracted from images in specific subfolders. It calculates a similarity score 
 and marks test cases as "Passed" or "Failed" based on the comparison. 
 
 The code also handles common subfolders between "Test Case Failed" and "Test Case Passed" 
 directories and ensures a clean comparison.'''



# Setting up stopwords
stop_words = set(stopwords.words('english'))

def remove_stopwords(text):
    """Remove stopwords from the given text."""
    return ' '.join([word for word in text.split() if word.lower() not in stop_words])

def get_column_index(sheet, target_header):
    """Retrieve the column index based on the header name."""
    for col_num, col_cells in enumerate(sheet.iter_cols(values_only=True)):
        if col_cells[0] == target_header:
            return col_num + 1  # 1-based index
    return None

def test_case_validation():
    # Define path to "Testcase" folder with Excel files
    testcase_folder_path = r"C:\Users\abdul\PycharmProjects\Automation_Optimum\Testcases"

    # Set base folder to "Screenshot"
    base_folder_path = r'C:\Users\abdul\PycharmProjects\Automation_Optimum\Screenshot'
    pass_folder_path = os.path.join(base_folder_path, "Test Case Passed")
    fail_folder_path = os.path.join(base_folder_path, "Test Case Failed")

    # Check for common sub-folders in 'Test Case Failed' and 'Test Case Passed'
    for sub_folder_name in os.listdir(fail_folder_path):
        fail_sub_folder_path = os.path.join(fail_folder_path, sub_folder_name)
        pass_sub_folder_path = os.path.join(pass_folder_path, sub_folder_name)

        if os.path.isdir(fail_sub_folder_path) and os.path.isdir(pass_sub_folder_path):
            # Check if both sub-folders have images
            fail_images = [f for f in os.listdir(fail_sub_folder_path) if f.endswith(('.png', '.jpg', '.jpeg'))]
            pass_images = [f for f in os.listdir(pass_sub_folder_path) if f.endswith(('.png', '.jpg', '.jpeg'))]

            if fail_images and pass_images:
                print(f"Found common sub-folder: {sub_folder_name}. Deleting from 'Test Case Passed'...")
                # Delete the sub-folder from 'Test Case Passed'
                for filename in os.listdir(pass_sub_folder_path):
                    os.remove(os.path.join(pass_sub_folder_path, filename))
                os.rmdir(pass_sub_folder_path)

    # Iterate over each Excel file in the "Testcase" folder
    for excel_file in os.listdir(testcase_folder_path):
        if not excel_file.startswith("~") and excel_file.endswith('.xlsx'):
            excel_path = os.path.join(testcase_folder_path, excel_file)
            print(f"Processing Excel file: {excel_file}")

            # Load the Excel file
            wb = load_workbook(filename=excel_path)
            ws = wb.active

            # Check for Expected Results column
            expected_results_col_index = get_column_index(ws, "Expected Results")
            if not expected_results_col_index:
                print(f"Skipping file {excel_file} due to missing 'Expected Results' column.")
                continue

            # Create "Actual Results" column if not present
            last_col = ws.max_column
            if "Actual Results" not in [cell.value for cell in ws[1]]:
                last_col += 1
                ws.cell(row=1, column=last_col).value = "Actual Results"

            # Create a dictionary mapping IDs to their row numbers
            id_to_rownum = {}
            for row_num, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), 2):
                id_to_rownum[row[0]] = row_num

            processed_ids = set()  # Keep track of the IDs you process from the 'Pass' folder

            # Iterate over sub-folders within the Pass folder
            for sub_folder_name in os.listdir(pass_folder_path):
                sub_folder_path = os.path.join(pass_folder_path, sub_folder_name)

                if os.path.isdir(sub_folder_path) and sub_folder_name in id_to_rownum:
                    # If an ID is found and processed, add it to the processed_ids set
                    processed_ids.add(sub_folder_name)
                    target_row = id_to_rownum[sub_folder_name]
                    expected_text = ws.cell(row=target_row, column=expected_results_col_index).value.lower()


                    # Extract text from images in the subfolder and concatenate
                    all_texts = []
                    for filename in os.listdir(sub_folder_path):
                        if filename.endswith(('.png', '.jpg', '.jpeg')):
                            img_path = os.path.join(sub_folder_path, filename)
                            text = extract_text_from_image(img_path).lower()
                            all_texts.append(text)

                    concatenated_text = ' '.join(all_texts)
                    match = re.search(r'"(.*?)"', concatenated_text)

                    if match:
                        concatenated_text = match.group(1)


                    # Remove stopwords from the expected text
                    filtered_expected_text = remove_stopwords(concatenated_text)

                    similarity_score = fuzz.partial_ratio(filtered_expected_text, concatenated_text)
                    print('similarity:', similarity_score)
                    print('Expected Result:', filtered_expected_text)
                    print('Image text:', concatenated_text)

                    if similarity_score >= 50:
                        ws.cell(row=target_row, column=last_col).value = "Passed"
                    else:
                        ws.cell(row=target_row, column=last_col).value = "Failed"

            # Check for IDs that weren't processed and mark them as "Failed"
            for test_id, row_num in id_to_rownum.items():
                if test_id not in processed_ids:
                    print(f"Warning: ID {test_id} from Excel file {excel_file} doesn't have a corresponding subfolder.")
                    ws.cell(row=row_num, column=last_col).value = "Failed"

            # Save changes to the Excel file.
            wb.save(excel_path)
            print(f"Processed and saved changes to {excel_file}.")

    print("All files have been processed.")


if __name__ == "__main__":
    # nltk.download('stopwords')  # Ensure the stopwords dataset is downloaded
    test_case_validation()


